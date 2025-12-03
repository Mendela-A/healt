import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# --- ГЛОБАЛЬНІ ЗМІННІ ---
# Шляхи до файлів тепер визначаються користувачем
FILE_PATH1 = None # Шлях до файлу-довідника (patients.xlsx)
df1 = None
df2 = None
changes = []
filtered_df = None  # Для зберігання фільтрованих даних на другій вкладці
main_df = None       # Основний файл для вкладки 3
compare_df = None    # Файл для порівняння (вкладка 3)
mismatches = []      # Список невідповідностей, які виводимо на вкладці 3
# Вибір колонок для порівняння
main_name_col = None
main_date_col = None
comp_name_col = None
comp_date_col = None

# --- ФУНКЦІЇ ДЛЯ ОБРОБКИ ДАНИХ ТА GUI ---

def get_sheet_names(file_path):
    """Отримує список назв аркушів з Excel-файлу."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        messagebox.showerror("Помилка", f"Не вдалося прочитати аркуші файлу {file_path}: {e}")
        return []

def load_first_file_dialog():
    """Відкриває діалог вибору першого файлу та відображає вибір аркуша."""
    global FILE_PATH1
    
    selected_path = filedialog.askopenfilename(
        title="Вибрати файл-довідник (patients.xlsx)",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    
    if not selected_path:
        return # Користувач скасував вибір

    FILE_PATH1 = selected_path
    sheet_names = get_sheet_names(FILE_PATH1)
    
    if sheet_names:
        show_sheet_selection_dialog(sheet_names)
    else:
        FILE_PATH1 = None # Скидаємо шлях, якщо аркуші не завантажилися
        info_label.config(text="Помилка завантаження аркушів. Спробуйте інший файл.")


def show_sheet_selection_dialog(sheet_names):
    """Створює нове вікно для вибору аркуша."""
    
    sheet_window = tk.Toplevel(root)
    sheet_window.title("Вибір аркуша")
    sheet_window.geometry("300x150")
    sheet_window.transient(root) # Щоб вікно було поверх головного

    tk.Label(sheet_window, text="Оберіть аркуш для завантаження:", pady=10).pack()

    # Змінна для зберігання обраного значення
    selected_sheet = tk.StringVar(sheet_window)
    selected_sheet.set(sheet_names[0]) # Встановлюємо перший аркуш як дефолтний

    # Комбобокс для вибору аркуша
    sheet_chooser = ttk.Combobox(sheet_window, textvariable=selected_sheet, values=sheet_names, state="readonly")
    sheet_chooser.pack(pady=5)
    
    # Кнопка для підтвердження вибору
    def confirm_selection():
        sheet_name = selected_sheet.get()
        sheet_window.destroy()
        load_first_file_data(sheet_name) # Завантажуємо дані з обраного аркуша

    tk.Button(sheet_window, text="Завантажити", command=confirm_selection, 
              bg="#4CAF50", fg="white").pack(pady=10)


def load_first_file_data(sheet_name):
    """Завантажує дані з обраного аркуша першого файлу."""
    global df1
    
    try:
        df1 = pd.read_excel(FILE_PATH1, sheet_name=sheet_name)
        df1 = df1[['his_num', 'name']] # Вибираємо потрібні колонки
        
        info_label.config(text=f"Файл-довідник ('{FILE_PATH1.split('/')[-1]}', Аркуш: '{sheet_name}') завантажено.")
        # Активуємо кнопку завантаження другого файлу
        load_button.config(state=tk.NORMAL)
    except Exception as e:
        df1 = None
        messagebox.showerror("Помилка завантаження", f"Помилка при завантаженні даних з {FILE_PATH1} ({sheet_name}): {e}")
        load_button.config(state=tk.DISABLED)


def load_and_process_files():
    """Викликається після вибору другого файлу. Обробляє та відображає дані."""
    global df2, changes
    
    if df1 is None:
        messagebox.showwarning("Попередження", "Спочатку успішно завантажте файл-довідник.")
        return
        
    file_path2 = filedialog.askopenfilename(
        title="Вибрати файл для оновлення (hospitalizations_11.xlsx)",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    
    if not file_path2:
        return 
        
    try:
        df2 = pd.read_excel(file_path2)
    except Exception as e:
        messagebox.showerror("Помилка завантаження", f"Помилка при завантаженні {file_path2}: {e}")
        return

    # 2. Обробка даних (як у оригінальному коді)
    # Перейменування колонок
    df2_renamed = df2.rename(columns={
        'Номер паперової історії хвороби': 'his_num',
        'ПІБ пацієнта': 'name'
    })

    # Об'єднання
    merged_df = df2_renamed.merge(
        df1[['his_num', 'name']], 
        on='his_num', 
        how='left', 
        suffixes=('_old', '_new')
    )

    # Пошук змін
    changes = []
    for idx, row in merged_df.iterrows():
        if pd.notna(row['name_new']) and row['name_old'] != row['name_new']:
            changes.append({
                'his_num': row['his_num'],
                'old_name': row['name_old'],
                'new_name': row['name_new'],
                'index': idx 
            })
            
    # 3. Оновлення GUI
    update_treeview()
    info_label.config(text=f"Файл '{file_path2.split('/')[-1]}' завантажено. Знайдено змін: {len(changes)}")
    save_button.config(state=tk.NORMAL)


def update_treeview():
    """Очищує та заповнює таблицю (Treeview) новими даними."""
    for item in tree.get_children():
        tree.delete(item)
        
    for i, change in enumerate(changes, 1):
        tree.insert('', tk.END, values=(
            i,
            change['his_num'],
            change['old_name'],
            change['new_name']
        ))

def load_hospitalization_file():
    """Завантажує файл госпіталізацій та відображає відфільтровані дані."""
    global filtered_df
    
    file_path = filedialog.askopenfilename(
        title="Вибрати файл госпіталізацій",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    
    if not file_path:
        return
    
    try:
        df = pd.read_excel(file_path)
        
        # Вибираємо потрібні колонки
        selected_columns = df[[
            'Номер паперової історії хвороби',
            'Дата госпіталізації',
            'Медпрацівник. відповідальний за епізод',
            'ПІБ пацієнта',
            'Дата та час виписки',
            'Помилка оплати НСЗУ'
        ]]
        
        # Фільтруємо дані - залишаємо тільки рядки, де 'Помилка оплати НСЗУ' не пуста
        filtered_df = selected_columns[selected_columns['Помилка оплати НСЗУ'].notnull()]
        
        # Оновлюємо таблицю на вкладці 2
        update_filtered_treeview()
        
        status_label.config(text=f"Завантажено! Знайдено {len(filtered_df)} записів з помилками оплати", fg="green")
        
    except KeyError as e:
        messagebox.showerror("Помилка", f"Файл не містить необхідної колонки: {e}")
    except Exception as e:
        messagebox.showerror("Помилка", f"Помилка при завантаженні файлу: {e}")

def update_filtered_treeview():
    """Оновлює таблицю з фільтрованими даними."""
    if filtered_df is None:
        return
    
    # Очищаємо таблицю
    for item in tree_filtered.get_children():
        tree_filtered.delete(item)
    
    # Заповнюємо таблицю
    for idx, row in filtered_df.iterrows():
        tree_filtered.insert('', tk.END, values=(
            idx + 1,
            row['Номер паперової історії хвороби'],
            row['Дата госпіталізації'],
            row['Медпрацівник. відповідальний за епізод'],
            row['ПІБ пацієнта'],
            row['Дата та час виписки'],
            row['Помилка оплати НСЗУ']
        ))

def save_changes():
    """Зберігає зміни в новий Excel-файл з виділенням."""
    if df2 is None or len(changes) == 0:
        messagebox.showinfo("Інформація", "Немає змін для збереження або не завантажено файл для оновлення.")
        return
    
    # ... (Логіка збереження і форматування залишається без змін) ...
    df3 = df2.copy()
    name_changes = {change['his_num']: change['new_name'] for change in changes}
    changed_indices = []
    
    for idx, row in df3.iterrows():
        his_num = row.get('Номер паперової історії хвороби')
        if his_num in name_changes:
            df3.at[idx, 'ПІБ пацієнта'] = name_changes[his_num]
            changed_indices.append(idx) 
    
    output_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile='hospitalizations_11_updated.xlsx',
        title="Зберегти оновлений файл як...",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )

    if not output_file:
        return 
        
    df3.to_excel(output_file, index=False)
    
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Помилка форматування", f"Не вдалося відкрити файл для форматування: {e}")
        return

    name_col_idx = None
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1), start=1):
        if col[0].value == 'ПІБ пацієнта':
            name_col_idx = idx
            break
    
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    if name_col_idx:
        for idx in changed_indices:
            cell = ws.cell(row=idx + 2, column=name_col_idx) 
            cell.fill = yellow_fill
    
    wb.save(output_file)
    
    messagebox.showinfo("Успіх", 
                        f"Файл '{output_file.split('/')[-1]}' успішно збережено!\n"
                        f"Змінено записів: {len(changed_indices)}")


# --- СТВОРЕННЯ GUI ---
root = tk.Tk()
root.title("Перегляд та оновлення ПІБ пацієнтів")
root.geometry("900x600")

# --- СТВОРЕННЯ ВКЛАДОК (NOTEBOOK) ---
notebook = ttk.Notebook(root)
notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

# ===== ВКЛАДКА 1: Основна функціональність =====
tab1 = ttk.Frame(notebook)
notebook.add(tab1, text="Оновлення довідника ПІБ")

title_label = tk.Label(tab1, text="Знайдені розбіжності в ПІБ пацієнтів", 
                        font=("Arial", 14, "bold"), pady=10)
title_label.pack()

info_label = tk.Label(tab1, text="1. Виберіть файл-довідник. 2. Виберіть файл для оновлення.", 
                        font=("Arial", 10), pady=5)
info_label.pack()

# --- КНОПКИ УПРАВЛІННЯ ---
button_frame = tk.Frame(tab1)
button_frame.pack(pady=10)

# 1. Кнопка вибору основного файлу
load_file1_button = tk.Button(button_frame, text="Вибрати файл-довідник", 
                              command=load_first_file_dialog, bg="#FF9800", fg="white",
                              font=("Arial", 10, "bold"), padx=10, pady=5)
load_file1_button.pack(side=tk.LEFT, padx=5)

# 2. Кнопка для завантаження другого файлу (заблокована до вибору першого)
load_button = tk.Button(button_frame, text="Вибрати файл H24", 
                        command=load_and_process_files, bg="#2196F3", fg="white",
                        font=("Arial", 10, "bold"), padx=10, pady=5, state=tk.DISABLED)
load_button.pack(side=tk.LEFT, padx=5)

# 3. Кнопка для збереження змін (заблокована до знаходження змін)
save_button = tk.Button(button_frame, text="Зберегти зміни", 
                        command=save_changes, bg="#4CAF50", fg="white",
                        font=("Arial", 10, "bold"), padx=10, pady=5, state=tk.DISABLED)
save_button.pack(side=tk.LEFT, padx=5)

exit_button = tk.Button(button_frame, text="Вихід", 
                        command=root.quit, bg="#f44336", fg="white",
                        font=("Arial", 10, "bold"), padx=10, pady=5)
exit_button.pack(side=tk.LEFT, padx=5)

# --- ТАБЛИЦЯ (TREEVIEW) ---
frame = ttk.Frame(tab1)
frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

scrollbar = ttk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

columns = ("№", "Номер історії", "Старе ПІБ", "Нове ПІБ")
tree = ttk.Treeview(frame, columns=columns, show='headings', yscrollcommand=scrollbar.set)
scrollbar.config(command=tree.yview)

tree.heading("№", text="№")
tree.heading("Номер історії", text="Номер історії")
tree.heading("Старе ПІБ", text="Старе ПІБ")
tree.heading("Нове ПІБ", text="Нове ПІБ")

tree.column("№", width=50, anchor='center')
tree.column("Номер історії", width=120, anchor='center')
tree.column("Старе ПІБ", width=300)
tree.column("Нове ПІБ", width=300)

tree.pack(fill=tk.BOTH, expand=True)

# ===== ВКЛАДКА 2: Фільтровані дані по помилкам оплати =====
tab2 = ttk.Frame(notebook)
notebook.add(tab2, text="Помилки оплати")

# Заголовок та кнопка завантаження файлу
header_frame = tk.Frame(tab2, bg="lightblue", padx=10, pady=10)
header_frame.pack(fill=tk.X)

title_label2 = tk.Label(header_frame, text="Записи з помилками оплати НСЗУ", 
                        font=("Arial", 12, "bold"), bg="lightblue")
title_label2.pack(side=tk.LEFT)

load_file_button2 = tk.Button(header_frame, text="Завантажити файл", 
                              command=load_hospitalization_file, bg="#4CAF50", fg="white",
                              font=("Arial", 10, "bold"), padx=15, pady=5)
load_file_button2.pack(side=tk.RIGHT, padx=5)

# Статус
status_label = tk.Label(tab2, text="Файл не завантажено", 
                        font=("Arial", 9), fg="#666")
status_label.pack(pady=5)

# Таблиця з фільтрованими даними
frame_filtered = ttk.Frame(tab2)
frame_filtered.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

scrollbar_filtered = ttk.Scrollbar(frame_filtered)
scrollbar_filtered.pack(side=tk.RIGHT, fill=tk.Y)

columns_filtered = ("№", "Номер історії", "Дата госпіталізації", "Медпрацівник", "ПІБ пацієнта", "Дата виписки", "Помилка оплати")
tree_filtered = ttk.Treeview(frame_filtered, columns=columns_filtered, show='headings', yscrollcommand=scrollbar_filtered.set)
scrollbar_filtered.config(command=tree_filtered.yview)

tree_filtered.heading("№", text="№")
tree_filtered.heading("Номер історії", text="Номер історії")
tree_filtered.heading("Дата госпіталізації", text="Дата госпіталізації")
tree_filtered.heading("Медпрацівник", text="Медпрацівник")
tree_filtered.heading("ПІБ пацієнта", text="ПІБ пацієнта")
tree_filtered.heading("Дата виписки", text="Дата виписки")
tree_filtered.heading("Помилка оплати", text="Помилка оплати")

tree_filtered.column("№", width=40, anchor='center')
tree_filtered.column("Номер історії", width=80, anchor='center')
tree_filtered.column("Дата госпіталізації", width=100, anchor='center')
tree_filtered.column("Медпрацівник", width=150)
tree_filtered.column("ПІБ пацієнта", width=150)
tree_filtered.column("Дата виписки", width=100, anchor='center')
tree_filtered.column("Помилка оплати", width=200)

tree_filtered.pack(fill=tk.BOTH, expand=True)

# ===== ВКЛАДКА 3: Порівняння записів =====
tab3 = ttk.Frame(notebook)
notebook.add(tab3, text="Порівняння записів")

controls_frame = tk.Frame(tab3, pady=8)
controls_frame.pack(fill=tk.X, padx=10)

load_main_btn = tk.Button(controls_frame, text="Завантажити основний файл", 
                          command=lambda: load_main_file_tab3(), bg="#2196F3", fg="white",
                          font=("Arial", 10, "bold"), padx=10, pady=5)
load_main_btn.pack(side=tk.LEFT, padx=5)

load_compare_btn = tk.Button(controls_frame, text="Завантажити файл для порівняння", 
                          command=lambda: load_compare_file_tab3(), bg="#FF9800", fg="white",
                          font=("Arial", 10, "bold"), padx=10, pady=5)
load_compare_btn.pack(side=tk.LEFT, padx=5)

compare_btn = tk.Button(controls_frame, text="Порівняти", 
                        command=lambda: run_compare_tab3(), bg="#4CAF50", fg="white",
                        font=("Arial", 10, "bold"), padx=10, pady=5, state=tk.DISABLED)
compare_btn.pack(side=tk.LEFT, padx=5)

diag_btn = tk.Button(controls_frame, text="Перевірити статус", 
                     command=lambda: show_diagnostics(), bg="#607D8B", fg="white",
                     font=("Arial", 9), padx=8, pady=4)
diag_btn.pack(side=tk.LEFT, padx=5)

status_main_label = tk.Label(controls_frame, text="Основний файл: не завантажено", font=("Arial", 9))
status_main_label.pack(side=tk.RIGHT, padx=5)

status_compare_label = tk.Label(controls_frame, text="Файл для порівняння: не завантажено", font=("Arial", 9))
status_compare_label.pack(side=tk.RIGHT, padx=5)

# --- Маппінг колонок (дві строки: окремо для основного файлу та для файлу порівняння)
main_map_frame = tk.Frame(tab3)
main_map_frame.pack(fill=tk.X, padx=10, pady=(6,2))

tk.Label(main_map_frame, text="Основний файл:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0,8))
tk.Label(main_map_frame, text="ПІБ:").pack(side=tk.LEFT, padx=(0,4))
main_name_var = tk.StringVar()
main_name_cb = ttk.Combobox(main_map_frame, textvariable=main_name_var, values=[], state='readonly', width=30)
main_name_cb.pack(side=tk.LEFT, padx=(0,12))

tk.Label(main_map_frame, text="Дата:").pack(side=tk.LEFT, padx=(0,4))
main_date_var = tk.StringVar()
main_date_cb = ttk.Combobox(main_map_frame, textvariable=main_date_var, values=[], state='readonly', width=18)
main_date_cb.pack(side=tk.LEFT, padx=(0,12))

tk.Label(main_map_frame, text="№ історії:").pack(side=tk.LEFT, padx=(0,4))
main_his_var = tk.StringVar()
main_his_cb = ttk.Combobox(main_map_frame, textvariable=main_his_var, values=[], state='readonly', width=20)
main_his_cb.pack(side=tk.LEFT, padx=(0,8))

comp_map_frame = tk.Frame(tab3)
comp_map_frame.pack(fill=tk.X, padx=10, pady=(2,10))

tk.Label(comp_map_frame, text="Файл для порівняння:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0,8))
tk.Label(comp_map_frame, text="ПІБ:").pack(side=tk.LEFT, padx=(0,4))
comp_name_var = tk.StringVar()
comp_name_cb = ttk.Combobox(comp_map_frame, textvariable=comp_name_var, values=[], state='readonly', width=30)
comp_name_cb.pack(side=tk.LEFT, padx=(0,12))

tk.Label(comp_map_frame, text="Дата:").pack(side=tk.LEFT, padx=(0,4))
comp_date_var = tk.StringVar()
comp_date_cb = ttk.Combobox(comp_map_frame, textvariable=comp_date_var, values=[], state='readonly', width=18)
comp_date_cb.pack(side=tk.LEFT, padx=(0,12))

# Таблиця для невідповідностей
tree_mismatch_frame = ttk.Frame(tab3)
tree_mismatch_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

scroll_m = ttk.Scrollbar(tree_mismatch_frame)
scroll_m.pack(side=tk.RIGHT, fill=tk.Y)

cols_m = ("№", "his_num", "name", "data")
tree_mismatch = ttk.Treeview(tree_mismatch_frame, columns=cols_m, show='headings', yscrollcommand=scroll_m.set)
scroll_m.config(command=tree_mismatch.yview)

tree_mismatch.heading("№", text="№")
tree_mismatch.heading("his_num", text="Номер історії")
tree_mismatch.heading("name", text="name")
tree_mismatch.heading("data", text="data")

tree_mismatch.column("№", width=50, anchor='center')
tree_mismatch.column("his_num", width=120, anchor='center')
tree_mismatch.column("name", width=260)
tree_mismatch.column("data", width=160, anchor='center')

tree_mismatch.pack(fill=tk.BOTH, expand=True)

# --- ДЕКЛАРАЦІЯ ФУНКЦІЙ ДЛЯ ВКЛАДКИ 3 ---
def load_main_file_tab3():
    global main_df
    path = filedialog.askopenfilename(title="Вибрати основний файл", filetypes=(("Excel files","*.xlsx"),("All files","*.*")))
    if not path:
        return

    # Отримати список аркушів
    sheets = get_sheet_names(path)
    if not sheets:
        # Якщо аркушів не знайдено, пробуємо завантажити всі дані
        try:
            main_df = pd.read_excel(path)
            status_main_label.config(text=f"Основний файл: {path.split('/')[-1]} ({len(main_df)})")
            try:
                cols = list(main_df.columns)
                main_name_cb['values'] = cols
                main_date_cb['values'] = cols
                main_his_cb['values'] = cols
                if 'name' in cols:
                    main_name_var.set('name')
                if 'data' in cols:
                    main_date_var.set('data')
                if 'his_num' in cols:
                    main_his_var.set('his_num')
                # try common Ukrainian header as fallback
                if not main_his_var.get() and 'Номер паперової історії хвороби' in cols:
                    main_his_var.set('Номер паперової історії хвороби')
            except Exception:
                pass
            # Перевірити чи можна увімкнути кнопку порівняння
            try:
                check_enable_compare()
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося завантажити основний файл: {e}")
        return

    # Показати вікно вибору аркуша
    sheet_win = tk.Toplevel(root)
    sheet_win.title("Вибір аркуша для основного файлу")
    sheet_win.geometry("320x140")
    sheet_win.transient(root)

    tk.Label(sheet_win, text="Оберіть аркуш:", pady=8).pack()
    sel_sheet = tk.StringVar(sheet_win)
    sel_sheet.set(sheets[0])
    cb = ttk.Combobox(sheet_win, textvariable=sel_sheet, values=sheets, state='readonly')
    cb.pack(pady=5)

    def confirm():
        sheet_name = sel_sheet.get()
        sheet_win.destroy()
        global main_df
        try:
            main_df = pd.read_excel(path, sheet_name=sheet_name)
            status_main_label.config(text=f"Основний файл: {path.split('/')[-1]} - {sheet_name} ({len(main_df)})")
            # Перевірити чи можна увімкнути кнопку порівняння
            try:
                check_enable_compare()
            except Exception:
                if compare_df is not None:
                    compare_btn.config(state=tk.NORMAL)
            # Наповнити списки колонок для маппінгу
            try:
                cols = list(main_df.columns)
                main_name_cb['values'] = cols
                main_date_cb['values'] = cols
                main_his_cb['values'] = cols
                # Спробуємо встановити дефолтні варіанти, якщо такі знайдені
                if 'name' in cols:
                    main_name_var.set('name')
                if 'data' in cols:
                    main_date_var.set('data')
                if 'his_num' in cols:
                    main_his_var.set('his_num')
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося завантажити аркуш '{sheet_name}': {e}")

    tk.Button(sheet_win, text="Завантажити", command=confirm, bg="#4CAF50", fg="white", padx=10, pady=5).pack(pady=10)

def load_compare_file_tab3():
    global compare_df
    path = filedialog.askopenfilename(title="Вибрати файл для порівняння", filetypes=(("Excel files","*.xlsx"),("All files","*.*")))
    if not path:
        return

    sheets = get_sheet_names(path)
    if not sheets:
        try:
            compare_df = pd.read_excel(path)
            status_compare_label.config(text=f"Файл для порівняння: {path.split('/')[-1]} ({len(compare_df)})")
            try:
                check_enable_compare()
            except Exception:
                if main_df is not None:
                    compare_btn.config(state=tk.NORMAL)
            try:
                cols = list(compare_df.columns)
                comp_name_cb['values'] = cols
                comp_date_cb['values'] = cols
                if 'ПІБ пацієнта' in cols:
                    comp_name_var.set('ПІБ пацієнта')
                if 'Дата закриття епізоду' in cols:
                    comp_date_var.set('Дата закриття епізоду')
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося завантажити файл для порівняння: {e}")
        return

    # Показати діалог вибору аркуша
    sheet_win = tk.Toplevel(root)
    sheet_win.title("Вибір аркуша для файлу порівняння")
    sheet_win.geometry("320x140")
    sheet_win.transient(root)

    tk.Label(sheet_win, text="Оберіть аркуш:", pady=8).pack()
    sel_sheet = tk.StringVar(sheet_win)
    sel_sheet.set(sheets[0])
    cb = ttk.Combobox(sheet_win, textvariable=sel_sheet, values=sheets, state='readonly')
    cb.pack(pady=5)

    def confirm():
        sheet_name = sel_sheet.get()
        sheet_win.destroy()
        global compare_df
        try:
            compare_df = pd.read_excel(path, sheet_name=sheet_name)
            status_compare_label.config(text=f"Файл для порівняння: {path.split('/')[-1]} - {sheet_name} ({len(compare_df)})")
            try:
                check_enable_compare()
            except Exception:
                if main_df is not None:
                    compare_btn.config(state=tk.NORMAL)
            # Наповнити списки колонок для маппінгу
            try:
                cols = list(compare_df.columns)
                comp_name_cb['values'] = cols
                comp_date_cb['values'] = cols
                # Спробуємо встановити дефолтні варіанти, якщо такі знайдені
                if 'ПІБ пацієнта' in cols:
                    comp_name_var.set('ПІБ пацієнта')
                if 'Дата закриття епізоду' in cols:
                    comp_date_var.set('Дата закриття епізоду')
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося завантажити аркуш '{sheet_name}': {e}")

    tk.Button(sheet_win, text="Завантажити", command=confirm, bg="#4CAF50", fg="white", padx=10, pady=5).pack(pady=10)

def check_enable_compare():
    """Увімкнути/вимкнути кнопку порівняння залежно від наявності обох файлів."""
    try:
        if main_df is not None and compare_df is not None:
            compare_btn.config(state=tk.NORMAL)
        else:
            compare_btn.config(state=tk.DISABLED)
    except Exception:
        pass

def show_diagnostics():
    """Показує діагностичну інформацію про стан вкладки порівняння."""
    try:
        main_loaded = main_df is not None
        comp_loaded = compare_df is not None
        main_rows = len(main_df) if main_loaded else 0
        comp_rows = len(compare_df) if comp_loaded else 0
        compare_state = compare_btn.cget('state')
        main_name = main_name_var.get() or '<не обрано>'
        main_date = main_date_var.get() or '<не обрано>'
        comp_name = comp_name_var.get() or '<не обрано>'
        comp_date = comp_date_var.get() or '<не обрано>'

        info = (
            f"Основний файл завантажено: {main_loaded}\n"
            f"Кількість рядків (основний): {main_rows}\n"
            f"Файл для порівняння завантажено: {comp_loaded}\n"
            f"Кількість рядків (порівняння): {comp_rows}\n\n"
            f"Стан кнопки 'Порівняти': {compare_state}\n\n"
            f"Вибрані колонки:\n"
            f"  Основний - ПІБ: {main_name}\n"
            f"  Основний - Дата: {main_date}\n"
            f"  Порівняння - ПІБ: {comp_name}\n"
            f"  Порівняння - Дата: {comp_date}\n"
        )

        messagebox.showinfo("Діагностика", info)
        print(info)
    except Exception as e:
        messagebox.showerror("Діагностика - помилка", f"Помилка при отриманні статусу: {e}")

def run_compare_tab3():
    """Порівнює записи: з основного файлу беремо 'name' та 'data',
    з файлу для порівняння беремо 'ПІБ пацієнта' та 'Дата закриття епізоду'.
    Виводимо записи з основного файлу, які не мають відповідника у файлі для порівняння.
    """
    global mismatches
    mismatches = []
    if main_df is None or compare_df is None:
        messagebox.showwarning("Попередження", "Завантажте обидва файли перед порівнянням.")
        return

    # Визначаємо, які колонки використовувати на основі вибору користувача (або дефолтів)
    c_name_col = comp_name_var.get() or 'ПІБ пацієнта'
    c_date_col = comp_date_var.get() or 'Дата закриття епізоду'
    try:
        comp = compare_df[[c_name_col, c_date_col]].copy()
    except Exception as e:
        cols = list(compare_df.columns) if compare_df is not None else []
        messagebox.showerror("Помилка", f"Файл для порівняння не містить вибраних колонок ({c_name_col}, {c_date_col}): {e}\n\nНаявні колонки: {cols}")
        return

    comp['name_norm'] = comp[c_name_col].astype(str).str.strip().str.lower()
    comp['date_norm'] = pd.to_datetime(comp[c_date_col], errors='coerce').dt.normalize()
    comp_set = set((r['name_norm'], r['date_norm']) for _, r in comp.iterrows())

    # Підготовка основного файлу
    m_name_col = main_name_var.get() or 'name'
    m_date_col = main_date_var.get() or 'data'
    try:
        # також беремо колонку з номером історії, якщо вона вказана
        m_his_col = main_his_var.get() or 'his_num'
        cols_to_take = [m_name_col, m_date_col]
        if m_his_col not in cols_to_take:
            cols_to_take.insert(0, m_his_col)
        main = main_df[cols_to_take].copy()
    except Exception as e:
        cols = list(main_df.columns) if main_df is not None else []
        messagebox.showerror("Помилка", f"Основний файл не містить вибраних колонок ({m_name_col}, {m_date_col}): {e}\n\nНаявні колонки: {cols}")
        return

    # Нормалізуємо колонки
    main['name_norm'] = main[m_name_col].astype(str).str.strip().str.lower()
    main['date_norm'] = pd.to_datetime(main[m_date_col], errors='coerce').dt.normalize()

    for idx, row in main.iterrows():
        key = (row['name_norm'], row['date_norm'])
        if key not in comp_set:
            his_val = ''
            # отримуємо оригінальне значення номера історії, якщо колонка є
            if m_his_col in main.columns:
                his_val = row.get(m_his_col, '')
            mismatches.append({'idx': idx, 'his_num': his_val, 'name': row[m_name_col], 'data': row[m_date_col]})

    update_mismatch_treeview()
    messagebox.showinfo("Результат", f"Порівняння завершено. Знайдено невідповідностей: {len(mismatches)}")

def update_mismatch_treeview():
    for it in tree_mismatch.get_children():
        tree_mismatch.delete(it)
    for i, row in enumerate(mismatches, 1):
        tree_mismatch.insert('', tk.END, values=(i, row.get('his_num', ''), row['name'], row['data']))

# --- ЗАПУСК GUI ---
root.mainloop()